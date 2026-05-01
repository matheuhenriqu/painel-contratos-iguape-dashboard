#!/usr/bin/env python3
"""Gera a camada estática de dados para o Painel de Contratos."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "ID",
    "Modalidade",
    "Nº Modalidade",
    "Objeto",
    "Processo",
    "Contrato",
    "Empresa Contratada",
    "Valor (R$)",
    "Valor (Descrição)",
    "Data Início",
    "Data Vencimento",
    "Dias p/ Vencimento",
    "Status",
    "Gestor",
    "Fiscal",
    "Observações",
]


FIELD_MAP = {
    "ID": "id",
    "Modalidade": "modalidade",
    "Nº Modalidade": "numeroModalidade",
    "Objeto": "objeto",
    "Processo": "processo",
    "Contrato": "contrato",
    "Empresa Contratada": "empresaContratada",
    "Valor (R$)": "valor",
    "Valor (Descrição)": "valorDescricao",
    "Data Início": "dataInicio",
    "Data Vencimento": "dataVencimento",
    "Dias p/ Vencimento": "diasPlanilha",
    "Status": "statusOriginal",
    "Gestor": "gestor",
    "Fiscal": "fiscal",
    "Observações": "observacoes",
}


def find_default_source() -> Path:
    candidates = [
        Path("contratos.xlsx"),
        Path.cwd().parent / "contratos.xlsx",
        Path.home() / "Documents" / "contratos.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        normalized = " ".join(value.strip().split())
        return normalized or None
    return value


def to_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return None


def read_contracts(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, data_only=True)
    worksheet = workbook["CONTRATOS"] if "CONTRATOS" in workbook.sheetnames else workbook.active

    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    missing_columns = [column for column in EXPECTED_COLUMNS if column not in headers]

    contracts: list[dict[str, Any]] = []
    for row_number in range(2, worksheet.max_row + 1):
        raw = {
            headers[column - 1]: worksheet.cell(row_number, column).value
            for column in range(1, worksheet.max_column + 1)
            if headers[column - 1]
        }

        if not any(value is not None and str(value).strip() for value in raw.values()):
            continue

        contract: dict[str, Any] = {"linhaOrigem": row_number}
        for original_name, field_name in FIELD_MAP.items():
            value = clean_value(raw.get(original_name))
            if field_name == "valor":
                value = to_number(value)
            contract[field_name] = value

        contracts.append(contract)

    return {
        "meta": {
            "source": source.name,
            "sourceModifiedAt": datetime.fromtimestamp(source.stat().st_mtime, timezone.utc).isoformat(),
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sheet": worksheet.title,
            "recordCount": len(contracts),
            "columns": headers,
            "missingExpectedColumns": missing_columns,
            "notes": [
                "Dados derivados da planilha original; a planilha não é alterada.",
                "Dias e status de vencimento são recalculados no navegador.",
            ],
        },
        "contracts": contracts,
    }


def write_outputs(payload: dict[str, Any], json_path: Path, js_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    js_path.parent.mkdir(parents=True, exist_ok=True)

    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    json_path.write_text(json_text + "\n", encoding="utf-8")
    js_path.write_text(f"window.CONTRATOS_DATA = {json_text};\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Converte contratos.xlsx para data/contratos.json e data/contratos.js.")
    parser.add_argument("--source", type=Path, default=find_default_source(), help="Caminho da planilha contratos.xlsx.")
    parser.add_argument("--json", type=Path, default=Path("data/contratos.json"), help="Arquivo JSON de saída.")
    parser.add_argument("--js", type=Path, default=Path("data/contratos.js"), help="Arquivo JS de saída para compatibilidade local.")
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {args.source}")

    payload = read_contracts(args.source)
    write_outputs(payload, args.json, args.js)
    print(f"{payload['meta']['recordCount']} contratos exportados para {args.json} e {args.js}")


if __name__ == "__main__":
    main()
